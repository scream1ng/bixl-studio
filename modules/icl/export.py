"""ICL Excel export — fills the real IXL Inspection Check List template.

Never mutates the template on disk; loads a fresh copy per call and streams the
result back as bytes. Template: docx/icl/Template - ICL.xlsx.
"""
from __future__ import annotations

import base64
import io
import os
from copy import copy
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage

_TEMPLATE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    "docx", "icl", "Template - ICL.xlsx",
)

# template layout (1-based)
_CHECK_START = 10          # first check row
_TEMPLATE_ROWS = 6         # rows 10..15 in the blank template
_IMG_BAND = 16             # the tall drawing-band row that follows the checks


def _fmt(v) -> str:
    try:
        return "%g" % round(float(v), 2)
    except (TypeError, ValueError):
        return ""


def _check_text(c: dict) -> str:
    """CHECKS cell, e.g. 'Check distance 34mm' / 'Check hole Ø12mm' / bend 90°."""
    desc = (c.get("desc") or "").strip()
    t = c.get("type") or "dist"
    if t == "visual":
        return desc
    vs = _fmt(c.get("value"))
    if not vs:
        return desc
    if t == "dia":
        dim = f"Ø{vs}mm"
    elif t == "rad":
        dim = f"R{vs}mm"
    elif t == "angle":
        dim = f"{vs}°"
    else:
        dim = f"{vs}mm"
    return f"{desc} {dim}" if desc else f"Check {dim}"


def _limit_text(c: dict) -> str:
    """Render the LIMITS cell: SOP ref for Reference gauge, else ' +/- 0.5mm'."""
    if (c.get("gauge") or "") == "Visual":
        return c.get("ref") or ""
    t = c.get("type") or "dist"
    if t == "visual":
        return ""
    tol = str(c.get("tol") or "").strip()
    if not tol:
        return ""
    unit = " deg" if t == "angle" else "mm"
    return f" +/- {tol}{unit}"


def _decode_image(data_url: str) -> bytes | None:
    if not data_url:
        return None
    b64 = data_url.split(",", 1)[1] if "," in data_url else data_url
    try:
        return base64.b64decode(b64)
    except Exception:
        return None


def _grow_check_rows(ws, extra: int) -> None:
    """Insert `extra` check rows before the drawing band, copying row-15 style.

    openpyxl moves cell values/styles on insert but NOT merged ranges, so we
    unmerge everything at/below the band, insert, then re-merge shifted.
    """
    moved = []
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= _IMG_BAND:
            moved.append((mc.min_col, mc.min_row, mc.max_col, mc.max_row))
            ws.unmerge_cells(str(mc))

    ws.insert_rows(_IMG_BAND, extra)

    for c1, r1, c2, r2 in moved:
        ws.merge_cells(start_row=r1 + extra, start_column=c1,
                       end_row=r2 + extra, end_column=c2)

    # style + merge the new check rows (copy the last template check row, 15)
    src = _CHECK_START + _TEMPLATE_ROWS - 1  # row 15
    src_h = ws.row_dimensions[src].height
    for i in range(extra):
        r = _CHECK_START + _TEMPLATE_ROWS + i  # 16, 17, ...
        if src_h:
            ws.row_dimensions[r].height = src_h
        for col in range(1, ws.max_column + 1):
            s, d = ws.cell(src, col), ws.cell(r, col)
            if s.has_style:
                d.font = copy(s.font)
                d.border = copy(s.border)
                d.fill = copy(s.fill)
                d.alignment = copy(s.alignment)
                d.number_format = s.number_format
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)   # C:D desc
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)   # E:F limits


def _col_px(ws, col: int) -> int:
    """Column width in px — WITHOUT materializing a ColumnDimension entry."""
    letter = get_column_letter(col)
    cd = ws.column_dimensions
    w = cd[letter].width if letter in cd else None
    if not w:
        w = ws.sheet_format.defaultColWidth or 8.43
    return int(round(w * 7 + 5))


# 0-based start columns per picture count (A=0, B=1, C=2, ...). Max 4.
_SLOT_START_BY_COUNT = {
    1: [4],            # E
    2: [2, 8],         # C, I
    3: [2, 4, 8],      # C, E, I
    4: [0, 3, 6, 12],  # A, D, G, M
}


def _place_images(ws, screenshots: list[dict], band_row: int) -> None:
    """Embed up to 4 screenshots; start columns depend on how many there are."""
    shots = screenshots[:4]
    slot_start = _SLOT_START_BY_COUNT.get(len(shots))
    if not slot_start:
        return
    total_px = sum(_col_px(ws, c) for c in range(1, ws.max_column + 1))
    starts = [sum(_col_px(ws, c) for c in range(1, sc + 1)) for sc in slot_start]
    band_px = (ws.row_dimensions[band_row].height or 200) * 96 / 72
    for i, shot in enumerate(shots):
        raw = _decode_image(shot.get("image", ""))
        if not raw:
            continue
        try:
            with PILImage.open(io.BytesIO(raw)) as pim:
                w, h = pim.size
        except Exception:
            continue
        slot_px = (starts[i + 1] - starts[i]) if i + 1 < len(starts) else (total_px - starts[i])
        scale = min(slot_px * 0.96 / w, band_px * 0.92 / h) if w and h else 1.0
        iw, ih = w * scale, h * scale
        row_off = max(0.0, (band_px - ih) / 2.0)    # center vertically in the band
        img = XLImage(io.BytesIO(raw))
        marker = AnchorMarker(col=slot_start[i], colOff=0,
                              row=band_row - 1, rowOff=pixels_to_EMU(row_off))
        img.anchor = OneCellAnchor(
            _from=marker,
            ext=XDRPositiveSize2D(pixels_to_EMU(iw), pixels_to_EMU(ih)),
        )
        ws.add_image(img)


def fill_icl(
    part_no: str = "",
    cust_no: str = "",
    part_desc: str = "",
    plant: str = "",
    checks: list[dict] | None = None,
    screenshots: list[dict] | None = None,
) -> bytes:
    """Fill the IXL ICL template and return .xlsx bytes.

    checks:      [{balloon, type, desc, value, tol, gauge}, ...]
    screenshots: [{type:'iso'|'view', name, image: dataURL}, ...]  (max 3)
    """
    checks = checks or []
    screenshots = screenshots or []

    wb = load_workbook(_TEMPLATE)
    ws = wb.active

    # header fields
    if part_no:
        ws["I6"] = part_no
    if cust_no:
        ws["C6"] = cust_no
    if part_desc:
        ws["N6"] = part_desc
    if plant:
        ws["F6"] = plant
    ws["C4"] = date.today()

    # colour "At Setup / Hourly" red in the inspection-frequency line (A7)
    _af = ws["A7"].font
    ws["A7"] = CellRichText(
        TextBlock(InlineFont(rFont=_af.name, sz=_af.size, b=_af.bold),
                  "Inspection Frequency : "),
        TextBlock(InlineFont(rFont=_af.name, sz=_af.size, b=True, color="FFCC0000"),
                  "At Setup / Hourly"),
    )

    # repurpose the "FIRST OFF APPR" column as the balloon number column
    ws["B8"] = "NO"
    ws["B9"] = None
    _f = ws["B8"].font
    ws["B8"].font = Font(name=_f.name, size=10, bold=_f.bold,
                         italic=_f.italic, color=_f.color)
    _a = ws["A10"].font
    ws["A10"].font = Font(name=_a.name, size=10, bold=_a.bold,
                          italic=_a.italic, color=_a.color)

    # P1 is a CustomDocProps() formula that resolves to #NAME? outside its
    # original doc — replace with the part number (or blank).
    ws["P1"] = part_no or ""

    # grow the check block if needed
    extra = max(0, len(checks) - _TEMPLATE_ROWS)
    if extra:
        _grow_check_rows(ws, extra)
    band_row = _IMG_BAND + extra

    # clear any pre-filled template check rows, then write ours
    last_check_row = _CHECK_START + max(len(checks), _TEMPLATE_ROWS) - 1
    for r in range(_CHECK_START, last_check_row + 1):
        for col in (2, 3, 5, 7):  # B balloon, C desc, E limits, G gauge
            ws.cell(r, col).value = None

    # balloon number cells: same font as CHECKS cell (C10), centered
    c10 = ws.cell(_CHECK_START, 3)
    b_font, b_nf = copy(c10.font), c10.number_format
    b_align = Alignment(horizontal="center", vertical=c10.alignment.vertical)
    for i, c in enumerate(checks):
        r = _CHECK_START + i
        bcell = ws.cell(r, 2)
        bcell.value = c.get("balloon") or ""
        bcell.font = copy(b_font)
        bcell.alignment = copy(b_align)
        bcell.number_format = b_nf
        ws.cell(r, 3).value = _check_text(c)
        ws.cell(r, 5).value = _limit_text(c)
        ws.cell(r, 7).value = c.get("gauge") or ""

    # drop the template's example drawings, keep only the top-left logo
    def _is_logo(im):
        f = getattr(im.anchor, "_from", None)
        return f is not None and f.row == 0 and f.col == 0
    ws._images = [im for im in ws._images if _is_logo(im)]

    # iso/overview first, then ballooned views
    ordered = sorted(screenshots, key=lambda s: 0 if s.get("type") == "iso" else 1)
    _place_images(ws, ordered, band_row)

    # name the sheet after the part number (Excel: <=31 chars, no []:*?/\)
    if part_no:
        safe = "".join(ch for ch in str(part_no) if ch not in '[]:*?/\\')[:31].strip()
        if safe:
            ws.title = safe

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
