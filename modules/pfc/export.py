"""PFC Excel export — fills the IXL Process Flow Chart template.

openpyxl cannot author flowchart autoshapes/connectors, but it embeds images
cleanly, so the flowchart is rendered client-side (Mermaid → canvas → PNG) and
dropped into the chart band of the template as a single picture. Header cells
are filled from the parsed BOM model.

Never mutates the template on disk; loads a fresh copy per call. Template:
docx/pfc/Template - PFC.xlsx (header + legend, part-specific shapes stripped).
"""
from __future__ import annotations

import base64
import io
import os
import re
from datetime import date

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage

_TEMPLATE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    "docx", "pfc", "Template - PFC.xlsx",
)

# chart band of the template (1-based): below the 3-row header, above the legend,
# within the A4-portrait print area (A1:X52 → cols A..X).
_CHART_TOP_ROW = 5
_CHART_BOT_ROW = 46
_CHART_FIRST_COL = 1
_CHART_LAST_COL = 24   # column X — right edge of the portrait print area


def _decode_image(data_url: str) -> bytes | None:
    if not data_url:
        return None
    b64 = data_url.split(",", 1)[1] if "," in data_url else data_url
    try:
        return base64.b64decode(b64)
    except Exception:
        return None


def _col_px(ws, col: int) -> int:
    letter = get_column_letter(col)
    cd = ws.column_dimensions
    w = cd[letter].width if letter in cd else None
    if not w:
        w = ws.sheet_format.defaultColWidth or 8.43
    return int(round(w * 7 + 5))


def _row_px(ws, row: int) -> int:
    h = ws.row_dimensions[row].height if row in ws.row_dimensions else None
    if not h:
        h = ws.sheet_format.defaultRowHeight or 15.0
    return int(round(h * 96 / 72))


def _place_chart(ws, png: bytes) -> None:
    """Embed the flowchart PNG, scaled to fit the chart band, centred."""
    try:
        with PILImage.open(io.BytesIO(png)) as pim:
            w, h = pim.size
    except Exception:
        return
    if not w or not h:
        return

    box_w = sum(_col_px(ws, c) for c in range(_CHART_FIRST_COL, _CHART_LAST_COL + 1))
    box_h = sum(_row_px(ws, r) for r in range(_CHART_TOP_ROW, _CHART_BOT_ROW + 1))

    scale = min(box_w * 0.98 / w, box_h * 0.98 / h)
    iw, ih = w * scale, h * scale
    col_off = max(0.0, (box_w - iw) / 2.0)
    row_off = max(0.0, (box_h - ih) / 2.0)

    img = XLImage(io.BytesIO(png))
    marker = AnchorMarker(
        col=_CHART_FIRST_COL - 1, colOff=pixels_to_EMU(col_off),
        row=_CHART_TOP_ROW - 1, rowOff=pixels_to_EMU(row_off),
    )
    img.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(pixels_to_EMU(iw), pixels_to_EMU(ih)),
    )
    ws.add_image(img)


def fill_pfc(model: dict, chart_png: str | bytes = b"") -> bytes:
    """Fill the PFC template from a parsed BOM model and return .xlsx bytes.

    chart_png: the rendered flowchart as a data URL / base64 string / raw bytes.
    """
    header = model.get("header") or {}
    wb = load_workbook(_TEMPLATE)
    ws = wb.active

    part_no = (header.get("part_no") or "").strip()
    part_name = (header.get("part_name") or "").strip()
    desc = (header.get("description") or "").strip()
    plant = (header.get("plant") or "").strip()

    if part_no:
        ws["G2"] = part_no
    pd = "    ".join(x for x in (part_name, desc) if x)
    if pd:
        ws["G3"] = pd
    if plant:
        ws["U1"] = plant
    ws["R2"] = date.today()

    png = chart_png if isinstance(chart_png, (bytes, bytearray)) else _decode_image(chart_png)
    if png:
        _place_chart(ws, png)

    # name the sheet after the part (Excel: <=31 chars, no []:*?/\)
    if part_no:
        safe = "".join(ch for ch in f"PFC {part_no}" if ch not in '[]:*?/\\')[:31].strip()
        if safe:
            ws.title = safe

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
