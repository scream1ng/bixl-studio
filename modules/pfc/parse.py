"""BOM → process-flow model.

Input: the "BOM calculation transactions including details" Excel export (one
indented production tree). Output: a flat, ordered list of operations plus the
header fields needed for the PFC sheet.

The transaction export is a depth-first flatten of the production tree. Columns
(1-based): 1 Type, 2 Level, 3 Item/Resource, 4 Description, 5 Operation,
6 Cost group, 7 Unit.

Tree shape that matters here:
  - `Production` at level 0          → the finished part (header) + final pack op
  - `BOM` rows (levels 1..N)         → one operation each (WIP sub-assemblies)
  - leaf rows (`Item`/`Setup`/`Process`/`Service`) attach to an operation by a
    simple level rule: a leaf at level X belongs to the operation node at level
    X-1 (its parent). This is what puts the M6 stud (level 3) onto the Weld
    operation (BOM node level 2) and the steel sheet (level 5) onto Laser
    (BOM node level 4).
  - operations are numbered by descending level: the deepest BOM node is the
    first physical operation (10), the shallowest is the last before pack.
  - `Surcharge` / overhead rows are ignored — they carry no process meaning.
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

# column indices (1-based) in the BOM transaction export
_TYPE, _LEVEL, _ITEM, _DESC, _OP, _CGROUP = 1, 2, 3, 4, 5, 6


def _op_name(desc: str) -> str:
    """Operation name from a BOM node description, e.g.
    'WAG-12H-0006323 - LASER' → 'Laser'. Falls back to the whole string."""
    if not desc:
        return ""
    tail = desc.rsplit("-", 1)[-1].strip()
    # only trust the suffix if it's a real word, not a leftover code chunk
    # (e.g. '...-0006323' → reject; '...- LASER' → 'Laser')
    if tail and any(c.isalpha() for c in tail) and not tail.replace(" ", "").isdigit():
        return tail.title()
    return ""   # caller resolves a fallback name


def _split_part(desc: str) -> tuple[str, str]:
    """'WAG-12H-0006323 -BRACKET-BODY HARN' → ('WAG-12H-0006323', 'BRACKET-BODY HARN')."""
    if not desc:
        return "", ""
    parts = desc.strip().split(None, 1)
    code = parts[0].strip()
    name = (parts[1] if len(parts) > 1 else "").lstrip("- ").strip()
    return code, name


def _new_op(level: int, op_no: int | None, name: str, code: str = "") -> dict:
    return {
        "op_no": op_no,          # 10, 20, ... or None for the final pack step
        "level": level,
        "code": code,            # WIP item number, e.g. 818245LBWE
        "name": name,
        "machines": [],          # MACH cost-group resources
        "labor": [],             # LAB cost-group resources
        "raws": [],              # [{item, desc}] raw materials consumed here
        "outsourced": False,     # True if a Service (VEND) row feeds it
        "op_field": "",          # Operation-column value of a child step (fallback name)
    }


def _add_unique(lst: list, value: str) -> None:
    v = (value or "").strip()
    if v and v not in lst:
        lst.append(v)


def parse_bom(xlsx_bytes: bytes) -> dict:
    """Parse a BOM transaction export into {header, operations[]}.

    Robust to any number of operations / levels — nothing is hard-coded to the
    818245 sample beyond the column layout of the export itself.
    """
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb.active

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[_TYPE - 1] in (None, ""):
            continue
        rows.append(r)
    wb.close()

    def cell(row, idx):
        return row[idx - 1] if idx - 1 < len(row) else None

    header = {"part_no": "", "part_name": "", "description": "", "plant": "Press Shop"}

    # final pack node (the Production row, level 0)
    pack = _new_op(0, None, "Pack")

    # operation nodes keyed by level (BOM rows). level 0 = pack.
    op_by_level: dict[int, dict] = {0: pack}

    for row in rows:
        rtype = (cell(row, _TYPE) or "").strip()
        level = cell(row, _LEVEL)
        level = int(level) if level is not None else 0
        item = str(cell(row, _ITEM) or "").strip()
        desc = str(cell(row, _DESC) or "").strip()
        opfield = str(cell(row, _OP) or "").strip()
        cgroup = (cell(row, _CGROUP) or "").strip().upper()

        if rtype == "Production":
            header["part_no"] = item
            code, name = _split_part(desc)
            header["part_name"], header["description"] = code, name
            pack["code"] = item
            continue

        if rtype == "BOM":
            op = _new_op(level, None, _op_name(desc), code=item)
            op_by_level[level] = op
            continue

        # leaf row → attach to the operation node one level up (its parent)
        parent = op_by_level.get(level - 1)
        if parent is None:
            parent = pack  # orphan safety net

        if rtype == "Item":
            parent["raws"].append({"item": item, "desc": desc})
        elif rtype in ("Setup", "Process"):
            if cgroup == "MACH":
                _add_unique(parent["machines"], desc)
            elif cgroup == "LAB":
                _add_unique(parent["labor"], desc)
            if not parent["op_field"] and opfield:
                parent["op_field"] = opfield   # fallback name source
        elif rtype == "Service":
            parent["outsourced"] = True
            _add_unique(parent["machines"], desc)
        # Surcharge / OVH and anything else: ignored

    # number the BOM operations by descending level (deepest = first op = 10)
    bom_ops = sorted(
        (op for lvl, op in op_by_level.items() if lvl > 0),
        key=lambda o: o["level"],
        reverse=True,
    )
    for i, op in enumerate(bom_ops):
        op["op_no"] = (i + 1) * 10
    # the finished-good node is the final (pack/label) operation — number it next
    pack["op_no"] = (len(bom_ops) + 1) * 10

    operations = bom_ops + [pack]  # pack is always last

    # resolve any operation left without a name from the description:
    # Operation-column value → first machine → "Operation N"
    for op in operations:
        if not op["name"]:
            fallback = op["op_field"] or (op["machines"][0] if op["machines"] else "")
            op["name"] = fallback.title() if fallback else (
                f"Operation {op['op_no']}" if op["op_no"] else "Process")

    return {"header": header, "operations": operations}
